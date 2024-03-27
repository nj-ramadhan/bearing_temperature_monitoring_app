from kivymd.app import MDApp
from kivymd.toast import toast
from kivymd.uix.datatables import MDDataTable
from kivymd.uix.boxlayout import MDBoxLayout
from kivymd.uix.filemanager import MDFileManager
from kivymd.uix.textfield import MDTextField
from kivymd.uix.label import MDLabel
from kivy.lang import Builder
from kivy.core.window import Window
from kivy.clock import Clock
from kivy.config import Config
from kivy.metrics import dp
from kivy.garden.matplotlib.backend_kivyagg import FigureCanvasKivyAgg
from kivy.properties import ObjectProperty
from kivy.properties import NumericProperty
from datetime import datetime
from pathlib import Path
import matplotlib.pyplot as plt
import matplotlib.colors as mcolors
from scipy.signal import find_peaks
import numpy as np
import os
import minimalmodbus
import time
import serial
from serial.tools import list_ports
import openpyxl

plt.style.use('bmh')

colors = {
    "Red": {
        "A200": "#EE2222",
        "A500": "#EE2222",
        "A700": "#EE2222",
    },
    "Blue": {
        "200": "#196BA5",
        "500": "#196BA5",
        "700": "#196BA5",
    },
    "Light": {
        "StatusBar": "E0E0E0",
        "AppBar": "#202020",
        "Background": "#FFFFFF",
        "CardsDialogs": "#EEEEEE",
        "FlatButtonDown": "#CCCCCC",
    },
    "Dark": {
        "StatusBar": "101010",
        "AppBar": "#E0E0E0",
        "Background": "#111111",
        "CardsDialogs": "#000000",
        "FlatButtonDown": "#333333",
    },
}

field_pos = [
    [0.242,0.790],[0.242,0.990],
    [0.262,0.790],[0.262,0.990],
    [0.290,0.790],[0.290,0.990],
    [0.310,0.790],[0.310,0.990],
    [0.338,0.790],[0.338,0.990],
    [0.358,0.790],[0.358,0.990],

    [0.438,0.790],[0.438,0.990],
    [0.458,0.790],[0.458,0.990],
    [0.532,0.790],[0.532,0.990],
    [0.552,0.790],[0.552,0.990],

    [0.648,0.790],[0.648,0.990],
    [0.668,0.790],[0.668,0.990],
    [0.742,0.790],[0.742,0.990],
    [0.762,0.790],[0.762,0.990],
    # [515,290],[515,600],
    # [630,290],[630,600],
    # [715,290],[715,600],
    # [1060,290],[1060,600],
    # [1150,290],[1150,600],
    # [1460,290],[1460,600],
    # [1550,290],[1550,600],
    # [1955,290],[1955,600],
    # [2045,290],[2045,600],
    # [2355,290],[2355,600],
    # [2445,290],[2445,600],

    # [165,730],[255,730],[565,730],[655,730],[1060,730],[1150,730],[1460,730],[1550,730],[1955,730],[2045,730],[2355,730],[2445,730],
    # [165,1040],[255,1040],[565,1040],[655,1040],[1060,1040],[1150,1040],[1460,1040],[1550,1040],[1955,1040],[2045,1040],[2355,1040],[2445,1040],

    # [165,1170],[255,1170],[565,1170],[655,1170],[1060,1170],[1150,1170],[1460,1170],[1550,1170],[1955,1170],[2045,1170],[2355,1170],[2445,1170],
    # [165,1480],[255,1480],[565,1480],[655,1480],[1060,1480],[1150,1480],[1460,1480],[1550,1480],[1955,1480],[2045,1480],[2355,1480],[2445,1480],

    # [165,1610],[255,1610],[565,1610],[655,1610],[1060,1610],[1150,1610],[1460,1610],[1550,1610],[1955,1610],[2045,1610],[2355,1610],[2445,1610],
    # [165,1920],[255,1920],[565,1920],[655,1920],[1060,1920],[1150,1920],[1460,1920],[1550,1920],[1955,1920],[2045,1920],[2355,1920],[2445,1920]
]

DEBUG = False

TEMP_OFFSET = 2.5412
TEMP_GAIN = 5.0 * 1000.0 #channge from A to mA with gain

BEARING_TEMP_DATA = 100
NUMBER_OF_BEARINGS = 100

BEARING_TEMP_MIN = 60.5

db_bearing_temps = np.zeros([100, 100])
arr_bearing_temps = np.zeros(100)
calc_bearing_temps = 0.0

flag_autosave_data = False
flag_autosave_graph = False
graph_state = 0

class ScreenSplash(MDBoxLayout):
    screen_manager = ObjectProperty(None)
    screen_standby = ObjectProperty(None)
    app_window = ObjectProperty(None)
    
    def __init__(self, **kwargs):
        super(ScreenSplash, self).__init__(**kwargs)
        Clock.schedule_interval(self.update_progress_bar, 0.01)

    def update_progress_bar(self, *args):
        if (self.ids.progress_bar.value + 1) < 100:
            raw_value = self.ids.progress_bar_label.text.split("[")[-1]
            value = raw_value[:-2]
            value = eval(value.strip())
            new_value = value + 1
            self.ids.progress_bar.value = new_value
            self.ids.progress_bar_label.text = "Loading.. [{:} %]".format(new_value)
        else:
            self.ids.progress_bar.value = 100
            self.ids.progress_bar_label.text = "Loading.. [{:} %]".format(100)
            time.sleep(0.5)
            self.screen_manager.current = "screen_standby"
            return False

class ScreenStandby(MDBoxLayout):
    screen_manager = ObjectProperty(None)

    def __init__(self, **kwargs):
        super(ScreenStandby, self).__init__(**kwargs)

    def screen_standby(self):
        self.screen_manager.current = 'screen_standby'

    def screen_data(self):
        self.screen_manager.current = 'screen_data'

    def screen_dashboard(self):
        self.screen_manager.current = 'screen_dashboard'

    def exec_shutdown(self): 
        toast("Shutting down system")
        os.system("shutdown /s /t 1") #for windows os
        # os.system("shutdown -h now") #for linux os

class ScreenData(MDBoxLayout):
    screen_manager = ObjectProperty(None)

    def __init__(self, **kwargs):
        super(ScreenData, self).__init__(**kwargs)
        self.file_manager = MDFileManager(exit_manager=self.exit_manager, select_path=self.select_path)
        Clock.schedule_once(self.delayed_init, 1)

    def delayed_init(self, dt):
        self.data_tables = MDDataTable(
            use_pagination=True,
            pagination_menu_pos="auto",
            rows_num=5,
            column_data=[("Bearing ", dp(20))]+[(f"T{i}", dp(10)) for i in range(1,101)]
        )

        self.ids.layout_tables.add_widget(self.data_tables)

        self.fig, self.ax = plt.subplots()
        self.fig.tight_layout()
        
        self.ax.set_xlabel("Data No.", fontsize=10)
        self.ax.set_ylabel("Temp. [C]", fontsize=10)

        self.ids.layout_graph.add_widget(FigureCanvasKivyAgg(self.fig))        

    def open_data(self):
        self.file_manager.show(os.path.expanduser(os.getcwd() + "\data"))  # output manager to the screen
        self.manager_open = True

    def select_path(self, path: str):
        try:
            self.exit_manager(path)
        except:
            toast("error select file path")

    def exit_manager(self, *args):
        global db_bearing_temps
        global arr_bearing_temps
        try:
            toast("opening data")
            dataframe = openpyxl.load_workbook(*args)
            dataframe_opened = dataframe.active

            # Iterate the loop to read the cell values
            
            # for row in range(1, dataframe_opened.max_row):
            #     for col in dataframe_opened.iter_cols(1, dataframe_opened.max_column):
            #         print(col[row].value)

            # db_bearing_temps = np.array(dataframe_opened)
            db_bearing_temps = np.array([[i.value for i in j] for j in dataframe_opened['B2':'CW101']])
            arr_bearing_temps = db_bearing_temps[0]
            # arr_bearing_temps = arr_bearing_temps[arr_bearing_temps != np.array(None)]
            # db_bearing_temps = db_bearing_temps[db_bearing_temps != np.array(None)]
            # db_bearing_temps = [d for d in db_bearing_temps if not(None in d )]

            # print(dataframe_opened)
            print(db_bearing_temps)
            print(arr_bearing_temps)

            # data_set = np.loadtxt(*args, delimiter="\t", encoding=None, skiprows=1)
            # data_base_process = data_set.T
            self.update_table()
            self.update_graph()

            self.manager_open = False
            self.file_manager.close()
                   
        except Exception as e:
            print("An exception occurred:", e)
            toast('error open file')
            self.manager_open = False
            self.file_manager.close()

    def update_table(self):           
        global db_bearing_temps
        global arr_bearing_temps
        numbers = np.arange(1,101)
        numbered_db = np.vstack((numbers,db_bearing_temps.T))
        print(numbered_db)

        try:
            self.data_tables.row_data = numbered_db.T.tolist()
        
        except Exception as e:
            print("An exception occurred:", e)

    def update_graph(self, bearing_num=1 ):           
        global db_bearing_temps
        global arr_bearing_temps
        try:
            arr_num = bearing_num - 1
            self.fig, self.ax = plt.subplots()
            self.fig.tight_layout()
            
            peaks, _ = find_peaks(db_bearing_temps[arr_num], height=BEARING_TEMP_MIN)
            arr_bearing_temps = db_bearing_temps[arr_num][db_bearing_temps[arr_num] != np.array(None)]
            
            self.ax.set_xlabel("n", fontsize=10)
            self.ax.set_ylabel("Temp. [C]", fontsize=10)
            self.ax.set_ylim(0, 100)
            self.ax.set_xlim(0, arr_bearing_temps.size)
            self.ax.plot(arr_bearing_temps)
            self.ax.plot(peaks, arr_bearing_temps[peaks], "x")
            self.ax.plot(np.zeros_like(arr_bearing_temps) + BEARING_TEMP_MIN, "--", color="gray")

            if arr_bearing_temps[peaks].size == 0:
                calc_bearing_temps = np.min(arr_bearing_temps)
            else:
                calc_bearing_temps = arr_bearing_temps[peaks][0]

            self.ids.label_bearing_temp.text = str(calc_bearing_temps)
            
            self.ids.layout_graph.clear_widgets()
            self.ids.layout_graph.add_widget(FigureCanvasKivyAgg(self.fig))
        
        except Exception as e:
            print("An exception occurred:", e)
            toast('error find peaks')
    
    def update_bearing_num(self):
        self.update_graph(int(self.ids.text_bearing_num.text))
        

    def sort_on_num(self, data):
        try:
            return zip(
                *sorted(
                    enumerate(data),
                    key=lambda l: l[0][0]
                )
            )
        except:
            toast("Error sorting data")
            
    def save_data(self):
        global db_bearing_temps
        global arr_bearing_temps
        try:
            name_file = "\data\\" + self.ids.input_file_name.text + ".xlsx"
            name_file_now = datetime.now().strftime("\data\%d_%m_%Y_%H_%M_%S.xlsx")
            cwd = os.getcwd()
            if self.ids.input_file_name.text == "":
                disk = cwd + name_file_now
            else:
                disk = cwd + name_file
            print(disk)
            with open(disk,"wb") as f:
                np.savetxt(f, db_bearing_temps.T, fmt="%.3f",delimiter="\t",header="Bearing No. \t Temperature [C]")
            print("sucessfully save data")
            toast("sucessfully save data")
        except:
            print("error saving data")
            toast("error saving data")
            
    def autosave_data(self):
        global db_bearing_temps

        try:
            now = datetime.now().strftime("/%d_%m_%Y_%H_%M_%S.raw")
            cwd = os.getcwd()
            disk = cwd + "\data\\" + now #for windows os
            with open(disk,"wb") as f:
                np.savetxt(f, db_bearing_temps.T, fmt="%.3f",delimiter="\t",header="Bearing No.  \t Temperature")
            print("sucessfully auto save data to Default Directory")
            toast("Sucessfully save data to The Default Directory")
        except:
            print("Error auto save data")
            toast("Error auto saving data")

    def screen_standby(self):
        self.screen_manager.current = 'screen_standby'

    def screen_data(self):
        self.screen_manager.current = 'screen_data'

    def screen_dashboard(self):
        self.screen_manager.current = 'screen_dashboard'

    def exec_shutdown(self): 
        toast("Shutting down system")
        os.system("shutdown /s /t 1") #for windows os
        # os.system("shutdown -h now") #for linux os

class ScreenDashboard(MDBoxLayout):
    def __init__(self, **kwargs):
        super(ScreenDashboard, self).__init__(**kwargs)
        Clock.schedule_once(self.delayed_init, 1)

    def delayed_init(self, dt):
        self.move_left()

    def move_left(self):
        global field_pos
        self.ids.background_image.source = 'asset/kereta_kiri.jpg'

        try:
            self.ids.layout_text_temps.clear_widgets()
            for i in range(1,29):
                field = MDLabel(id=f'T_{i}', 
                                text=f'{i}',
                                theme_text_color= 'Primary',
                                pos_hint= {'center_x': (field_pos[i-1][0]),'center_y': (field_pos[i-1][1])}
                )
                self.ids.layout_text_temps.add_widget(field)

        except Exception as e:
            print("An exception occurred:", e)
            toast('error open screen')

    def move_right(self):
        global field_pos
        self.ids.background_image.source = 'asset/kereta_kanan.jpg'

        try:
            self.ids.layout_text_temps.clear_widgets()
            for i in range(1,29):
                field = MDLabel(id=f'T_{i}', 
                                text=f'{i}',
                                theme_text_color= 'Error',
                                pos_hint= {'center_x': (field_pos[i-1][0]),'center_y': (field_pos[i-1][1])}
                )
                self.ids.layout_text_temps.add_widget(field)

        except Exception as e:
            print("An exception occurred:", e)
            toast('error open screen')

    def screen_standby(self):
        self.screen_manager.current = 'screen_standby'

    def screen_data(self):
        self.screen_manager.current = 'screen_data'

    def screen_dashboard(self):
        self.screen_manager.current = 'screen_dashboard'

    def exec_shutdown(self): 
        toast("Shutting down system")
        os.system("shutdown /s /t 1") #for windows os
        # os.system("shutdown -h now") #for linux os

class BearingTemperatureMonitoringApp(MDApp):
    def build(self):
        self.theme_cls.colors = colors
        self.theme_cls.primary_palette = "Blue"
        self.icon = "asset\logo_kai.png" #for windows os
        Window.fullscreen = 'auto'
        Window.borderless = True
        Window.allow_screensaver = True

        screen = Builder.load_file("main.kv")
        return screen

if __name__ == "__main__":
    BearingTemperatureMonitoringApp().run()
